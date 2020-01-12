from django.contrib import admin
from django import forms
from django.contrib import admin
from .models import Teacher
from django.db.models import Count
from student.models import Student, StudentImagesData
from django.utils import timezone
from course.models import Course, Schedule, Attendance, ScheduleImagesData
from django_admin_listfilter_dropdown.filters import DropdownFilter, RelatedDropdownFilter
from rangefilter.filter import DateRangeFilter, DateTimeRangeFilter
from import_export import resources
from import_export.admin import ImportExportModelAdmin
import csv
from django.http import HttpResponse
from django.utils.safestring import mark_safe
from datetime import datetime
from import_export.resources import Resource, DeclarativeMetaclass
import django
from django.conf import settings
from django.conf.urls import url
from django.contrib import admin, messages
from django.contrib.admin.models import ADDITION, CHANGE, DELETION, LogEntry
from django.contrib.auth import get_permission_codename
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse, HttpResponseRedirect
from django.template.response import TemplateResponse
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.utils.encoding import force_text
from django.utils.module_loading import import_string
from django.utils.translation import gettext_lazy as _
from django.views.decorators.http import require_POST
import functools
import logging
import tablib
import traceback
from collections import OrderedDict
from copy import deepcopy
from diff_match_patch import diff_match_patch
import django
from django.conf import settings
from django.core.exceptions import ImproperlyConfigured, ValidationError
from django.core.management.color import no_style
from django.db import DEFAULT_DB_ALIAS, connections
from django.db.models.fields.related import ForeignObjectRel
from django.db.models.query import QuerySet
from django.db.transaction import (
    TransactionManagementError,
    atomic,
    savepoint,
    savepoint_commit,
    savepoint_rollback
)
import openpyxl
from django.utils.encoding import force_text
from django.utils.safestring import mark_safe
from course.resources import CoursesResource
from import_export.resources import ModelResource
from student.models import Student
from student.admin import StudentAdmin
import json
from django.db.models import Q, F


# Register your models here.
class TeacherChoiceField(forms.ModelChoiceField):
    @staticmethod
    def label_from_instance(obj):
        return "{0}: {1} {2}".format(obj.teacher_code, obj.first_name, obj.last_name)


class CourseAdmin(ImportExportModelAdmin):
    list_display = (
        'course_code', 'course_name', 'children_display', 'start_day', 'end_day', 'teacher', 'student_count',
        'day_of_week', 'time_start_of_course', 'duaration',)
    search_fields = ('course_code', 'course_name',)
    fieldsets = (
        (None, {
            'fields': ('course_code', 'course_name', 'start_day', 'end_day', 'teacher', 'day_of_week',
                       'time_start_of_course', 'time_duration',)
        }),
        ('Advance options', {
            'fields': ('students',),
            'description': 'option advance',
            'classes': ('collapse',),
        }),
    )
    readonly_fields = []
    filter_horizontal = ('students',)
    list_filter = (
        ('teacher', RelatedDropdownFilter),
        ('start_day', DateRangeFilter), ('end_day', DateTimeRangeFilter),
    )
    resource_class = CoursesResource

    list_per_page = 20
    date_hierarchy = 'start_day'
    raw_id_fields = ["teacher", ]

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        queryset = queryset.annotate(
            student_count=Count("students", distinct=True),
        )

        if request.user.is_superuser:
            queryset = Course.objects.all().annotate(student_count=Count("students", distinct=True), )
        else:
            try:
                queryset = Course.objects.filter(teacher=request.user.username).annotate(
                    student_count=Count("students", distinct=True), )
            except:
                queryset = Course.objects.none().annotate(student_count=Count("students", distinct=True), )
        return queryset

    def get_readonly_fields(self, request, obj=None):
        if obj:  # editing an existing object
            return self.readonly_fields + ['course_code', ]
        return self.readonly_fields

    @staticmethod
    def duaration(obj):
        # return obj.students.count()
        return " " + str(obj.time_duration) + " tiết "

    duaration.short_description = "Time duration"

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'teacher':
            return TeacherChoiceField(queryset=Teacher.objects.all())
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def children_display(self, obj):
        return ", ".join([
            # students.student_code for students in obj.students.all()
            students.first_name for students in obj.students.all()
        ])

    children_display.short_description = "Students List"

    @staticmethod
    def student_count(obj):
        # return obj.students.count()
        return obj.student_count

    student_count.admin_order_field = 'student_count'

    actions = ["export_as_csv_by_course", "export_as_csv_by_student"]

    @staticmethod
    def set_csv_file_name():
        return 'Courses'

    def export_as_csv_by_course(self, request, queryset):
        # noinspection PyProtectedMember
        meta = self.model._meta
        field_names = [field.name for field in meta.fields]
        name = self.set_csv_file_name()
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename={0}.csv'.format(name)
        file_name = 'filename={0}.csv'.format(name)
        response.write(u'\ufeff'.encode('utf8'))
        writer = csv.writer(response)

        writer.writerow(['Thông tin môn học'])
        writer.writerow(
            ['Mã môn học', 'Tên môn học', 'Phòng học', 'Ngày bắt đầu', 'Ngày kết thúc', 'Giáo viên giản dạy',
             'Ngày học trong tuần (Thứ)', 'Tiết bắt đầu (Tiết)', 'Thời gian học (Số tiết học)'])
        # writer.writerow(field_names)
        for obj in queryset:
            row = writer.writerow([getattr(obj, field) for field in field_names])
            student_count = Course.objects.all().filter(course_code=getattr(obj, 'course_code')).values('students').count()
            total = Schedule.objects.all().filter(course__course_code=getattr(obj, 'course_code')).count()
            str_total = "Tổng số buổi: " + str(total) + " Tổng số sinh viên: " + str(student_count)
            writer.writerow([str_total])
            writer.writerow("")
            writer.writerow(['Tên môn học', 'Mã môn học', 'Mã sinh viên', 'Tên sinh viên', 'Ngày học', 'Buổi học', 'Vắng'])

            temp = Attendance.objects.filter(
                schedule_code__course__course_code=getattr(obj, 'course_code')).all().values(
                'schedule_code__course__course_code',
                'schedule_code__course__course_name',
                'schedule_code__schedule_date',
                'schedule_code__schedule_number_of_day',
                'student__first_name',
                'student__student_code',
                'absent_status'
            ).values('schedule_code__course__course_name',
                     'schedule_code__course__course_code',
                     'student__student_code',
                     'student__first_name',
                     'schedule_code__schedule_date',
                     'schedule_code__schedule_number_of_day',
                     'absent_status',
                     )
            for data in temp:
                writer.writerow(data.values())

        return response

    export_as_csv_by_course.short_description = "Export Selected course by course"

    def export_as_csv_by_student(self, request, queryset):
        # noinspection PyProtectedMember
        meta = self.model._meta
        field_names = [field.name for field in meta.fields]
        name = self.set_csv_file_name()
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename={0}.csv'.format(name)
        file_name = 'filename={0}.csv'.format(name)
        response.write(u'\ufeff'.encode('utf8'))
        writer = csv.writer(response)

        writer.writerow(['Thông tin môn học'])
        writer.writerow(
            ['Mã môn học', 'Tên môn học', 'Phòng học', 'Ngày bắt đầu', 'Ngày kết thúc', 'Giáo viên giản dạy',
             'Ngày học trong tuần (Thứ)', 'Tiết bắt đầu (Tiết)', 'Thời gian học (Số tiết học)'])
        # writer.writerow(field_names)
        for obj in queryset:
            row = writer.writerow([getattr(obj, field) for field in field_names])
            total = Schedule.objects.all().filter(course__course_code=getattr(obj, 'course_code')).count()
            str_total = "Tổng số buổi: " + str(total)
            writer.writerow([str_total])
            writer.writerow(['Tên môn học', 'Mã môn học', 'Mã sinh viên', 'Tên sinh viên', 'Vắng'])

            temp = Attendance.objects.filter(
                schedule_code__course__course_code=getattr(obj, 'course_code')).all().values(
                'schedule_code__course__course_code',
                'schedule_code__course__course_name',
                'student__first_name',
                'student__student_code',
                'absent_status'
            ).values('schedule_code__course__course_name',
                     'schedule_code__course__course_code',
                     'student__student_code',
                     'student__first_name',
                     ).annotate(
                total=Count('absent_status', filter=Q(absent_status=False))).order_by()
            for data in temp:
                writer.writerow(data.values())
            # print(Attendance.objects.all().filter(schedule_code__course__course_code=getattr(obj,
            # 'course_code')).values( 'schedule_code').annotate(total=Count('student')).order_by('total'))
            print()
        return response

    export_as_csv_by_student.short_description = "Export Selected course by student"

    def write_to_tmp_storage(self, import_file, input_format):
        tmp_storage = self.get_tmp_storage_class()()
        data = bytes()
        for chunk in import_file.chunks():
            data += chunk

        tmp_storage.save(data, input_format.get_read_mode())
        return tmp_storage

    def import_action(self, request, *args, **kwargs):
        if not self.has_import_permission(request):
            raise PermissionDenied

        context = self.get_import_context_data()

        import_formats = self.get_import_formats()
        form_type = self.get_import_form()
        form_kwargs = self.get_form_kwargs(form_type, *args, **kwargs)
        form = form_type(import_formats,
                         request.POST or None,
                         request.FILES or None,
                         **form_kwargs)

        if request.POST and form.is_valid():
            input_format = import_formats[
                int(form.cleaned_data['input_format'])
            ]()
            # input_format = Excel
            import_file = form.cleaned_data['import_file']

            # first always write the uploaded file to disk as it may be a

            wb = openpyxl.load_workbook(import_file)
            worksheet = wb.active
            try:
                t_course_code = worksheet.cell(row=6, column=4).value
                t_course_code = t_course_code[5:]
                t_course_name = worksheet.cell(row=6, column=1).value
                t_course_name = t_course_name[8:]
                t_start_day = ''
                t_end_day = ''
                t_teacher = worksheet.cell(row=7, column=4).value
                t_teacher = t_teacher[15:]
                t_students = ""
                t_class_time = '2'
                t_class_time_calendar = '2'
                t_class_time_begin_time = '1'
            except Exception as e:
                return HttpResponse(
                    _(u"<h1>%s encountered while trying to read file: %s</h1>"
                      u"<h1> Make sure its a real import format</h1> " %
                      (type(e).__name__, import_file.name)))
            # create new wb for save data

            count_data_row = worksheet.max_row
            for x in range(10, 10 + count_data_row):
                if worksheet.cell(row=x, column=2).value:
                    t_students = t_students + "," + str(worksheet.cell(row=x, column=2).value)

            wb.create_sheet(title="Sheet1", index=0)
            us_ws = wb.active
            us_ws.append(["course_code", "course_name", "start_day", "end_day", "teacher", "students", "class_time",
                          "class_time_calendar", "class_time_begin_time"])
            # add row here
            us_ws.append([t_course_code, t_course_name, t_start_day, t_end_day, t_teacher, t_students, t_class_time,
                          t_class_time_calendar, t_class_time_begin_time])

            wb.save(import_file)
            tmp_storage = self.write_to_tmp_storage(import_file, input_format)
            try:
                data = tmp_storage.read(input_format.get_read_mode())
                if not input_format.is_binary() and self.from_encoding:
                    data = force_text(data, self.from_encoding)
                dataset = input_format.create_dataset(data)
                print(dataset)
            except UnicodeDecodeError as e:
                return HttpResponse(_(u"<h1>Imported file has a wrong encoding: %s</h1>" % e))
            except Exception as e:
                return HttpResponse(
                    _(u"<h1>%s encountered while trying to read file: %s</h1>" % (type(e).__name__, import_file.name)))
            res_kwargs = self.get_import_resource_kwargs(request, form=form, *args, **kwargs)
            resource = self.get_import_resource_class()(**res_kwargs)
            imp_kwargs = self.get_import_data_kwargs(request, form=form, *args, **kwargs)
            result = resource.import_data(dataset, dry_run=True,
                                          raise_errors=False,
                                          file_name=import_file.name,
                                          user=request.user,
                                          **imp_kwargs)

            context['result'] = result

            if not result.has_errors() and not result.has_validation_errors():
                initial = {
                    'import_file_name': tmp_storage.name,
                    'original_file_name': import_file.name,
                    'input_format': form.cleaned_data['input_format'],
                }
                confirm_form = self.get_confirm_import_form()
                initial = self.get_form_kwargs(form=form, **initial)
                context['confirm_form'] = confirm_form(initial=initial)
        else:
            res_kwargs = self.get_import_resource_kwargs(request, form=form, *args, **kwargs)
            resource = self.get_import_resource_class()(**res_kwargs)

        context.update(self.admin_site.each_context(request))

        context['title'] = _("Import")
        context['form'] = form
        context['opts'] = self.model._meta
        context['fields'] = [f.column_name for f in resource.get_user_visible_fields()]
        request.current_app = self.admin_site.name
        return TemplateResponse(request, [self.import_template_name],
                                context)


class CourseChoiceField(forms.ModelChoiceField):
    @staticmethod
    def label_from_instance(obj):
        return "{0}: {1} {2}".format(obj.course_code, obj.course_name, obj.teacher)


class ScheduleImagesDataInline(admin.TabularInline):
    model = ScheduleImagesData
    fields = ('schedule', 'image_data'),
    extra = 1
    # classes = 'collapse',


class ScheduleAdmin(admin.ModelAdmin):
    list_display = ('schedule_code', 'course_info', 'schedule_date', 'schedule_number_of_day')
    search_fields = ('schedule_code',)
    inlines = (ScheduleImagesDataInline,)
    fieldsets = (
        (None, {
            'fields': ('course', 'schedule_number_of_day')
        }),
    )
    search_fields = ('schedule_code', 'schedule_number_of_day')
    readonly_fields = ['schedule_code', 'course', 'schedule_date', 'schedule_number_of_day']
    # filter_horizontal = ('students',)
    list_filter = (
        ('course', RelatedDropdownFilter),
        ('schedule_date', DateRangeFilter),
    )

    list_per_page = 20
    raw_id_fields = ["course", ]

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        if request.user.is_superuser:
            queryset = Schedule.objects.all()
        else:
            try:
                queryset = Schedule.objects.filter(course__teacher__username=request.user.username)
            except:
                queryset = Course.objects.none()
        return queryset

    @staticmethod
    def course_info(obj):
        return obj.course.course_name + " - " + obj.course.course_code + ""

    course_info.short_description = 'Course'

    def get_readonly_fields(self, request, obj=None):
        if obj:  # editing an existing object
            return self.readonly_fields + ['schedule_code', 'course', 'schedule_date', 'schedule_number_of_day']
        return self.readonly_fields

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'course':
            return CourseChoiceField(queryset=Course.objects.all())
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


class ScheduleChoiceField(forms.ModelChoiceField):
    @staticmethod
    def label_from_instance(obj):
        return "{0}: {1} {2}".format(obj.course, obj.schedule_code, obj.schedule_number_of_day)


class AttendanceChoiceField(forms.ModelChoiceField):
    @staticmethod
    def label_from_instance(obj):
        return "{0}: {1}".format(obj.student_code, obj.first_name)


class AttendanceAdmin(admin.ModelAdmin):
    list_display = ('course_info', 'attendance_code', 'student_info', 'schedule_code', 'date', 'absent_status',
                    'image_data',)
    search_fields = ('attendance_code', 'absent_status')

    fieldsets = (
        (None, {
            'fields': ('student', 'schedule_code', 'absent_status', 'image_data',)
        }),

    )
    # filter_horizontal = ('schedule_code',)
    list_filter = (
        ('student', RelatedDropdownFilter),
        ('schedule_code', RelatedDropdownFilter),
    )

    list_per_page = 20
    raw_id_fields = ["schedule_code", "student"]

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        if request.user.is_superuser:
            queryset = Attendance.objects.all()
        else:
            try:
                queryset = Attendance.objects.filter(schedule_code__course__teacher__username=request.user.username)
            except:
                queryset = Attendance.objects.none()
        return queryset

    # def get_readonly_fields(self, request, obj=None):
    #   if obj:  # editing an existing object
    #       return self.readonly_fields + ('student_info', 'field2')
    #   return self.readonly_fields

    @staticmethod
    def student_info(obj):
        return obj.student.first_name

    student_info.short_description = 'Student'

    @staticmethod
    def course_info(obj):
        course_code = (Schedule.objects.filter(schedule_code=obj.schedule_code).values_list('course__course_name'))[0]
        return course_code

    course_info.short_description = 'Course'

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'student':
            return AttendanceChoiceField(queryset=Student.objects.all())
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


admin.site.register(Attendance, AttendanceAdmin)
admin.site.register(Schedule, ScheduleAdmin)
admin.site.register(Course, CourseAdmin)
