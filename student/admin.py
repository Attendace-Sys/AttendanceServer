from django.contrib import admin
from .models import Student
from django.db.models import Count
from student.models import Student, StudentImagesData
from django.utils import timezone
from course.models import Course
from django_admin_listfilter_dropdown.filters import DropdownFilter, RelatedDropdownFilter
from rangefilter.filter import DateRangeFilter, DateTimeRangeFilter
from import_export import resources
from import_export.admin import ImportExportModelAdmin
from student.resources import StudentsResource
import csv
from django.http import HttpResponse
from django.utils.safestring import mark_safe
from datetime import datetime
from import_export.resources import Resource, DeclarativeMetaclass, ModelResource
import django
from django.conf import settings
from django.conf.urls import url
from django.contrib import admin, messages
from django.contrib.admin.models import ADDITION, CHANGE, DELETION, LogEntry
from django.contrib.auth import get_permission_codename
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse, HttpResponseRedirect
from django.template.response import TemplateResponse
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.utils.encoding import force_text
from django.utils.module_loading import import_string
from django.utils.translation import gettext_lazy as _
from django.views.decorators.http import require_POST
import functools
import logging
import tablib
import traceback
from collections import OrderedDict
from copy import deepcopy
from diff_match_patch import diff_match_patch
import django
from django.conf import settings
from django.core.exceptions import ImproperlyConfigured, ValidationError
from django.core.management.color import no_style
from django.db import DEFAULT_DB_ALIAS, connections
from django.db.models.fields.related import ForeignObjectRel
from django.db.models.query import QuerySet
from django.db.transaction import (
    TransactionManagementError,
    atomic,
    savepoint,
    savepoint_commit,
    savepoint_rollback
)
import openpyxl
from django.utils.encoding import force_text
from django.utils.safestring import mark_safe
from django.contrib.admin.widgets import AdminFileWidget
from django.db import models



class ImageInline(admin.TabularInline):
    model = StudentImagesData
    fields = ('student', 'image_data'),
    extra = 1
    classes = 'collapse',


class StudentAdmin(ImportExportModelAdmin,):
    labels = {
        'first_name': 'Full Name'
    }
    list_display = (
        'student_code', 'get_full_name', 'email', 'username',)
    readonly_fields = []
    list_filter = ('student_code',)
    search_fields = ('student_code',)
    inlines = (ImageInline,)
    fieldsets = (
        (None, {
            'fields': ('student_code', 'password', 'first_name', 'email', 'student_video_data', 'comment',)
        }),
    )

    add_fieldsets = (
        (None, {
            # 'classes': ('wide',),
            'fields': ('student_code', 'first_name', 'email', 'password1', 'password2',),
        }),
    )

    def get_readonly_fields(self, request, obj=None):
        if obj:  # editing an existing object
            return self.readonly_fields + ['student_code', 'username', 'password']
        return self.readonly_fields

    def get_full_name(self, obj):
        return obj.get_full_name()

    get_full_name.short_description = 'Full name'

    resource_class = StudentsResource

    @staticmethod
    def student_image_show(student):
        return mark_safe('<img src="{url}" width="{width}" height={height} />'.format(
            url=student.student_image.url,
            # width=teacher.teacher_image.width,
            width=80,
            # height=teacher.teacher_image.height,
            height=80
        )
        )

    @staticmethod
    def student_full_image_show(student):
        return mark_safe('<img src="{url}" width="{width}" height={height} />'.format(
            url=student.student_image.url,
            width=student.student_image.width,
            height=student.student_image.height,
        )
        )

    student_full_image_show.short_description = 'Full Image'

    def write_to_tmp_storage(self, import_file, input_format):
        tmp_storage = self.get_tmp_storage_class()()
        data = bytes()
        for chunk in import_file.chunks():
            data += chunk

        tmp_storage.save(data, input_format.get_read_mode())
        return tmp_storage

    def import_action(self, request, *args, **kwargs):
        if not self.has_import_permission(request):
            raise PermissionDenied

        context = self.get_import_context_data()

        import_formats = self.get_import_formats()

        form_type = self.get_import_form()
        form_kwargs = self.get_form_kwargs(form_type, *args, **kwargs)
        form = form_type(import_formats,
                         request.POST or None,
                         request.FILES or None,
                         **form_kwargs)

        if request.POST and form.is_valid():
            input_format = import_formats[
                int(form.cleaned_data['input_format'])
            ]()
            import_file = form.cleaned_data['import_file']
            """
            code change data file excel here
            """
            wb = openpyxl.load_workbook(import_file)
            worksheet = wb.active
            worksheet.delete_rows(0, 9)
            worksheet.delete_cols(1, 1)
            worksheet.delete_cols(7, 6)
            # worksheet.append(["student_code", "first_name", "last_name", "email", "username", "password"])
            worksheet.insert_rows(1)
            begrow = 2
            endrow = worksheet.max_row
            for row in range(begrow, endrow + 1):  # just an example
                if worksheet.cell(row, 1).value is None:
                    worksheet.delete_rows(row, row + 1)
            wb.save(import_file)
            # check
            # first always write the uploaded file to disk as it may be a
            # memory file or else based on settings upload handlers
            tmp_storage = self.write_to_tmp_storage(import_file, input_format)

            # then read the file, using the proper format-specific mode
            # warning, big files may exceed memory
            try:
                data = tmp_storage.read(input_format.get_read_mode())
                if not input_format.is_binary() and self.from_encoding:
                    data = force_text(data, self.from_encoding)
                dataset = input_format.create_dataset(data)
                dataset.headers = ['student_code', 'first_name', 'email', 'username', 'password',
                                   'comment']
            except UnicodeDecodeError as e:
                return HttpResponse(_(u"<h1>Imported file has a wrong encoding: %s</h1>" % e))
            except Exception as e:
                return HttpResponse(
                    _(u"<h1>%s encountered while trying to read file: %s Try to import another file</h1>" % (
                        type(e).__name__, import_file.name)))

            # prepare kwargs for import data, if needed
            res_kwargs = self.get_import_resource_kwargs(request, form=form, *args, **kwargs)
            resource = self.get_import_resource_class()(**res_kwargs)

            # prepare additional kwargs for import_data, if needed
            imp_kwargs = self.get_import_data_kwargs(request, form=form, *args, **kwargs)

            result = resource.import_data(dataset, dry_run=True,
                                          raise_errors=False,
                                          file_name=import_file.name,
                                          user=request.user,
                                          **imp_kwargs)

            context['result'] = result

            if not result.has_errors() and not result.has_validation_errors():
                initial = {
                    'import_file_name': tmp_storage.name,
                    'original_file_name': import_file.name,
                    'input_format': form.cleaned_data['input_format'],
                }
                confirm_form = self.get_confirm_import_form()
                initial = self.get_form_kwargs(form=form, **initial)
                context['confirm_form'] = confirm_form(initial=initial)
        else:
            res_kwargs = self.get_import_resource_kwargs(request, form=form, *args, **kwargs)
            resource = self.get_import_resource_class()(**res_kwargs)

        context.update(self.admin_site.each_context(request))

        context['title'] = _("Import")
        context['form'] = form
        context['opts'] = self.model._meta
        # context['fields'] = [f.column_name for f in resource.get_user_visible_fields()]
        context['fields'] = ['student_code', 'first_name', 'email', 'username', 'password',
                             'comment']
        request.current_app = self.admin_site.name
        return TemplateResponse(request, [self.import_template_name],
                                context)


# Register your models Student.
admin.site.register(Student, StudentAdmin)
